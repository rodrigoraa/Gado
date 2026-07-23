from __future__ import annotations

from collections.abc import Iterable, Sequence
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from typing import NotRequired, TypedDict

from django.http import HttpResponse
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class SecaoRelatorio(TypedDict):
    titulo: str
    cabecalhos: Sequence[str]
    linhas: Sequence[Sequence[object]]
    resumo: NotRequired[str]


def _valor_planilha(valor: object) -> object:
    """Mantém números como números no XLSX e normaliza valores de apresentação."""

    if valor is None:
        return ""
    if isinstance(valor, (Decimal, date, datetime, int)) and not isinstance(valor, bool):
        return valor
    return str(valor)


def _nome_planilha(titulo: str, usados: set[str]) -> str:
    nome = "".join("-" if caractere in "[]:*?/\\" else caractere for caractere in titulo).strip()
    nome = (nome or "Relatório")[:31]
    base = nome
    contador = 2
    while nome.casefold() in usados:
        sufixo = f" {contador}"
        nome = f"{base[: 31 - len(sufixo)]}{sufixo}"
        contador += 1
    usados.add(nome.casefold())
    return nome


def _preencher_planilha(planilha: Worksheet, secao: SecaoRelatorio) -> None:
    planilha.append(list(secao["cabecalhos"]))
    preenchimento = PatternFill("solid", fgColor="215446")
    for celula in planilha[1]:
        celula.font = Font(color="FFFFFF", bold=True)
        celula.fill = preenchimento
        celula.alignment = Alignment(horizontal="center", vertical="center")
    for linha in secao["linhas"]:
        planilha.append([_valor_planilha(valor) for valor in linha])
    planilha.freeze_panes = "A2"
    planilha.auto_filter.ref = planilha.dimensions
    for coluna in range(1, len(secao["cabecalhos"]) + 1):
        maior = max(
            (
                len(str(planilha.cell(linha, coluna).value or ""))
                for linha in range(1, planilha.max_row + 1)
            ),
            default=10,
        )
        planilha.column_dimensions[get_column_letter(coluna)].width = min(max(maior + 2, 12), 45)


def gerar_xlsx_secoes(titulo: str, secoes: Sequence[SecaoRelatorio]) -> bytes:
    workbook = Workbook()
    planilha_inicial = workbook.active
    workbook.remove(planilha_inicial)
    usados: set[str] = set()
    for secao in secoes:
        planilha = workbook.create_sheet(_nome_planilha(secao["titulo"], usados))
        _preencher_planilha(planilha, secao)
    if not workbook.worksheets:
        workbook.create_sheet(_nome_planilha(titulo, usados))
    saida = BytesIO()
    workbook.save(saida)
    return saida.getvalue()


def gerar_xlsx(titulo: str, cabecalhos: Sequence[str], linhas: Iterable[Sequence[object]]) -> bytes:
    """API histórica usada pelo comando mensal; delega para o exportador multisseção."""

    secao: SecaoRelatorio = {
        "titulo": titulo,
        "cabecalhos": cabecalhos,
        "linhas": list(linhas),
    }
    return gerar_xlsx_secoes(titulo, [secao])


def resposta_xlsx(
    nome: str,
    titulo: str,
    cabecalhos: Sequence[str] | None = None,
    linhas: Sequence[Sequence[object]] | None = None,
    *,
    secoes: Sequence[SecaoRelatorio] | None = None,
) -> HttpResponse:
    if secoes is None:
        secoes = [
            {
                "titulo": titulo,
                "cabecalhos": cabecalhos or (),
                "linhas": linhas or (),
            }
        ]
    response = HttpResponse(
        gerar_xlsx_secoes(titulo, secoes),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{nome}.xlsx"'
    return response


def resposta_pdf(
    nome: str,
    titulo: str,
    cabecalhos: Sequence[str] | None = None,
    linhas: Sequence[Sequence[object]] | None = None,
    contexto: dict[str, object] | None = None,
    *,
    secoes: Sequence[SecaoRelatorio] | None = None,
) -> HttpResponse:
    from weasyprint import HTML

    if secoes is None:
        secoes = [
            {
                "titulo": titulo,
                "cabecalhos": cabecalhos or (),
                "linhas": linhas or (),
            }
        ]
    html = render_to_string(
        "relatorios/pdf.html",
        {"titulo": titulo, "secoes": secoes, **(contexto or {})},
    )
    pdf = HTML(string=html).write_pdf()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nome}.pdf"'
    return response
